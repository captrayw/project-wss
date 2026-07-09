"""
Excel round-trip for the year-by-year Data Inputs table.

`build_template(fe)` writes a pre-filled, colour-coded workbook (cream = fill me, grey = auto);
`parse_template(bytes, fe)` reads a filled workbook back and overlays ONLY the editable cells onto
the current frontend-shaped inputs, returning the merged dict + a count of cells updated.

Both share `_layout()` / `_editable()` so the writer and the parser can never disagree about which
cell is which, or which cells are inputs vs engine-derived. Values are stored in MODEL units
(fractions for %); percentage cells just carry an Excel number-format so the user sees "5.00%".

Editable windows (by year, matching InputPanel.tsx):
  * GDP / inflation / direct-mode expenditure : model start … baseline + 5  ('hard')
  * exchange rate / population / households    : model start … baseline − 1  ('hist')
  * service-level shares                       : start year and baseline year only ('svc')
Everything else in the table is an engine projection and is written grey / left blank.
"""
import io
import copy
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

EDIT_FILL = PatternFill('solid', fgColor='FFF9E6')   # cream — a cell to fill in
LOCK_FILL = PatternFill('solid', fgColor='F1F3F5')   # grey  — engine-calculated, leave blank
HEAD_FILL = PatternFill('solid', fgColor='1E3A5F')   # header band
HEADER_ROW = 5


def _layout(fe):
    """Return (rows, mode). Each row: (key, label, section, field, kind, number_format)."""
    cc = fe.get('country_config', {}) or {}
    ws_names = [cc.get(f'ws_serv{i+1}_name') or f'Level {i+1}' for i in range(5)]
    sn_names = [cc.get(f'san_serv{i+1}_name') or f'Level {i+1}' for i in range(5)]
    mode = ((fe.get('macro', {}) or {}).get('budget_input_mode')
            or (fe.get('bau', {}) or {}).get('budget_input_mode') or 'pct_gdp')
    rows = [
        ('macro.gdp_nominal_usd', 'Nominal GDP (US$ billion)', 'macro', 'gdp_nominal_usd', 'hard', '#,##0.000'),
        ('macro.inflation_nepal', 'Local inflation', 'macro', 'inflation_nepal', 'hard', '0.00%'),
        ('macro.inflation_us', 'US inflation', 'macro', 'inflation_us', 'hard', '0.00%'),
        ('macro.exchange_rate', 'Exchange rate (local per US$)', 'macro', 'exchange_rate', 'hist', '#,##0.0000'),
        ('population.pop_ts', 'Population (millions)', 'population', 'pop_ts', 'hist', '#,##0.000000'),
        ('population.hh_ts', 'Households (millions)', 'population', 'hh_ts', 'hist', '#,##0.000000'),
    ]
    for i in range(5):
        rows.append((f'water_service.serv{i+1}_ts', f'Water — % {ws_names[i]}', 'water_service', f'serv{i+1}_ts', 'svc', '0.00%'))
    for i in range(5):
        rows.append((f'sanitation_service.sserv{i+1}_ts', f'Sanitation — % {sn_names[i]}', 'sanitation_service', f'sserv{i+1}_ts', 'svc', '0.00%'))
    if mode == 'direct':
        rows.append(('bau.ws_expend_ts', 'Water actual expenditure (millions, real)', 'bau', 'ws_expend_ts', 'hard', '#,##0.00'))
        rows.append(('bau.san_expend_ts', 'Sanitation actual expenditure (millions, real)', 'bau', 'san_expend_ts', 'hard', '#,##0.00'))
    return rows, mode


def _editable(kind, i, bi, hard):
    """Is column index i (year offset from model start) an editable INPUT for this row kind?"""
    if kind == 'hard':
        return i <= hard          # historical + baseline + 5 forecast years
    if kind == 'hist':
        return i < bi             # historical actuals only (through baseline − 1)
    if kind == 'svc':
        return i == 0 or i == bi  # start year + baseline year only
    return False


def _span(fe):
    per = fe.get('period', {}) or {}
    start = int(per.get('model_start_year', 2011))
    baseline = int(per.get('baseline_year', 2025))
    end = int(per.get('forecast_end_year', 2040))
    if end < start:
        end = start
    return start, baseline, end, baseline - start, min(baseline + 5, end) - start


def build_template(fe: dict) -> io.BytesIO:
    start, baseline, end, bi, hard = _span(fe)
    years = list(range(start, end + 1))
    rows, mode = _layout(fe)
    cc = fe.get('country_config', {}) or {}

    wb = Workbook()
    ws = wb.active
    ws.title = 'WSS Inputs'

    ws['B1'] = 'WSS year-by-year data — input template'
    ws['B1'].font = Font(bold=True, size=14, color='1E3A5F')
    ws['B2'] = (f"{cc.get('country', '')} — {cc.get('area', '')}   |   currency: {cc.get('currency', 'LCU')}"
                f"   |   budget mode: {mode}")
    ws['B2'].font = Font(color='475569')
    ws['B3'] = ('Fill only the CREAM cells, then upload this file. Grey cells are auto-calculated by the '
                'model — leave them blank. Keep the full decimal precision shown; enter percentages as shown (e.g. 5.00%).')
    ws['B3'].font = Font(italic=True, color='B45309')

    ws.cell(HEADER_ROW, 1, 'key')
    hb = ws.cell(HEADER_ROW, 2, 'Indicator')
    hb.font = Font(bold=True, color='FFFFFF'); hb.fill = HEAD_FILL
    for c, yr in enumerate(years):
        hc = ws.cell(HEADER_ROW, 3 + c, yr)
        hc.font = Font(bold=True, color='FFFFFF'); hc.fill = HEAD_FILL
        hc.alignment = Alignment(horizontal='center')

    r = HEADER_ROW + 1
    for key, label, section, field, kind, numfmt in rows:
        ws.cell(r, 1, key)
        lc = ws.cell(r, 2, label)
        lc.font = Font(bold=True, color='1E3A5F')
        arr = (fe.get(section, {}) or {}).get(field, []) or []
        for i, yr in enumerate(years):
            cell = ws.cell(r, 3 + i)
            if _editable(kind, i, bi, hard):
                if i < len(arr) and arr[i] is not None:
                    cell.value = arr[i]
                cell.fill = EDIT_FILL
                cell.number_format = numfmt
            else:
                cell.fill = LOCK_FILL
        r += 1

    ws.column_dimensions['A'].hidden = True     # machine key column, hidden from the user
    ws.column_dimensions['B'].width = 34
    for c in range(len(years)):
        ws.column_dimensions[get_column_letter(3 + c)].width = 11
    ws.freeze_panes = ws.cell(HEADER_ROW + 1, 3)

    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    return out


def parse_template(file_bytes: bytes, fe: dict):
    """Overlay the filled template's editable cells onto `fe`. Returns (merged_inputs, cells_updated)."""
    merged = copy.deepcopy(fe)
    start, baseline, end, bi, hard = _span(merged)
    n = end - start + 1
    rows, _mode = _layout(merged)
    keymap = {key: (section, field, kind) for key, _l, section, field, kind, _f in rows}

    wb = load_workbook(io.BytesIO(file_bytes), data_only=True)
    ws = wb['WSS Inputs'] if 'WSS Inputs' in wb.sheetnames else wb.active

    # Locate the header row (col B == 'Indicator'), then map each year to its column.
    header_row = None
    for rr in range(1, min(ws.max_row, 40) + 1):
        if str(ws.cell(rr, 2).value).strip().lower() == 'indicator':
            header_row = rr
            break
    if header_row is None:
        raise ValueError('Could not find the template header row (expected an "Indicator" column). '
                         'Please upload the file downloaded from this tool, unmodified in structure.')

    year_col = {}
    for cc in range(3, ws.max_column + 1):
        v = ws.cell(header_row, cc).value
        try:
            year_col[int(v)] = cc
        except (TypeError, ValueError):
            continue

    changed = 0
    for rr in range(header_row + 1, ws.max_row + 1):
        key = ws.cell(rr, 1).value
        if key not in keymap:
            continue
        section, field, kind = keymap[key]
        arr = list((merged.get(section, {}) or {}).get(field, []) or [])
        for i in range(n):
            if not _editable(kind, i, bi, hard) or i >= len(arr):
                continue                       # keep array length; never zero-pad hard-value tails
            col = year_col.get(start + i)
            if not col:
                continue
            v = ws.cell(rr, col).value
            if v is None or v == '':
                continue                       # blank = keep the existing value
            try:
                arr[i] = float(v)
            except (TypeError, ValueError):
                continue
            changed += 1
        merged.setdefault(section, {})[field] = arr

    return merged, changed
